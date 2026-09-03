from pathlib import Path
import json
import os
import sys
import tempfile

import pymupdf
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import analizador_universal as analyzer
import reportes_profesionales as professional


def check(name, condition):
    if not condition:
        print("FAIL", name)
        raise AssertionError(name)
    print("PASS", name)


prompt = """Genera dashboard HTML, PDF y Excel analítico con ventas, toneladas, costo, utilidad y flete.
No calcules flete si no existe una regla empresarial aprobada. Incluye provenance y métricas bloqueadas."""

print("\n=== R10.18A E2E CROSS-FORMAT GOVERNED DELIVERABLES ===")
with tempfile.TemporaryDirectory() as td:
    root = Path(td)
    source = root / "ventas.csv"
    source.write_text(
        "Fecha,Cliente,Articulo,Toneladas_Vendidas,Importe_Venta,Costo,Refer,Vendedor,Almacen,Costo_Flete_Corto,Costo_Flete_Largo,Costo_Flete_Traspaso\n"
        "2026-01-10,Cliente A,Producto 1,10,1000,800,F1,Ana,Norte,10,20,5\n"
        "2026-02-10,Cliente B,Producto 2,20,2400,1800,F2,Beto,Sur,20,30,10\n",
        encoding="utf-8",
    )
    reports = root / "Reportes"
    reports.mkdir()
    old_reports = analyzer.base.REPORTES
    old_llm = os.environ.get("IA_DYNAMIC_DASHBOARD_LLM")
    analyzer.base.REPORTES = reports
    os.environ["IA_DYNAMIC_DASHBOARD_LLM"] = "0"
    try:
        result = analyzer.analyze_file(source, prompt)
    finally:
        analyzer.base.REPORTES = old_reports
        if old_llm is None:
            os.environ.pop("IA_DYNAMIC_DASHBOARD_LLM", None)
        else:
            os.environ["IA_DYNAMIC_DASHBOARD_LLM"] = old_llm

    check("analysis_ok", result["ok"] is True)
    paths = {kind: reports / result[kind] for kind in ("html", "excel", "pdf")}
    check("all_formats_created", all(path.is_file() and path.stat().st_size > 0 for path in paths.values()))

    html = paths["html"].read_text(encoding="utf-8", errors="replace")
    check("html_manifest", '"enterprise_deliverable_manifest"' in html and '"schema_version":"r10.18a"' in html)
    check("html_freight_blocked", '"id":"kpi:freight"' in html and '"status":"BLOCKED"' in html)

    workbook = load_workbook(paths["excel"], read_only=True, data_only=False)
    check("excel_governance_sheets", "Gobernanza" in workbook.sheetnames and "Capacidades" in workbook.sheetnames)
    capability_values = [str(cell.value or "") for row in workbook["Capacidades"].iter_rows() for cell in row]
    check("excel_freight_blocked", "kpi:freight" in capability_values and "BLOCKED" in capability_values)
    governance_values = [str(cell.value or "") for row in workbook["Gobernanza"].iter_rows() for cell in row]
    check("excel_manifest_version", "r10.18a" in governance_values)
    workbook.close()

    with pymupdf.open(paths["pdf"]) as document:
        pdf_text = "\n".join(page.get_text() for page in document)
    check("pdf_governance_page", "Gobernanza y trazabilidad" in pdf_text)
    check("pdf_freight_blocked", "kpi:freight" in pdf_text and "BLOCKED" in pdf_text)

    data_start = html.index("const DATA=") + len("const DATA=")
    data_end = html.index(";\n(()=>", data_start)
    manifest = json.loads(html[data_start:data_end])["plan"]["enterprise_deliverable_manifest"]
    generic = pd.DataFrame([{"Categoría":"A","Valor":1},{"Categoría":"B","Valor":2}])
    generic_profile = {
        "archivo":"generic.csv", "filas":2, "columnas":["Categoría","Valor"],
        "roles_detectados":{}, "deliverable_manifest":manifest,
    }
    generic_sections = {
        "KPIs_Generales":pd.DataFrame([
            ["Filas",2],["Columnas",2],["Celdas nulas %",0.0],["Filas duplicadas",0],
        ], columns=["Indicador","Valor"]),
        "Resultado":generic,
    }
    generic_xlsx = reports / "professional.xlsx"
    generic_pdf = reports / "professional.pdf"
    professional.excel_report_professional(generic_xlsx,prompt,generic_profile,{"type":"overview"},generic_sections,[],"",generic,generic,{},"general")
    professional.pdf_report_professional(generic_pdf,prompt,generic_profile,generic_sections,[],"","general")
    generic_workbook = load_workbook(generic_xlsx,read_only=True)
    check("professional_excel_manifest", "Gobernanza" in generic_workbook.sheetnames and "Capacidades" in generic_workbook.sheetnames)
    generic_workbook.close()
    with pymupdf.open(generic_pdf) as document:
        generic_pdf_text = "\n".join(page.get_text() for page in document)
    check("professional_pdf_manifest", "Gobernanza y trazabilidad" in generic_pdf_text and "kpi:freight" in generic_pdf_text)

print("PASS R10.18A E2E CROSS-FORMAT GOVERNED DELIVERABLES")
