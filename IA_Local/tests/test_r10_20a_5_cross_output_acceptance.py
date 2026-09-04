from pathlib import Path
import json
import sys
import tempfile

import fitz
import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from dashboard_dynamic import build_dashboard_plan, generate_dynamic_dashboard
from enterprise_design_system import DESIGN_SYSTEM_VERSION
import reportes_profesionales as professional


def check(name, condition):
    if not condition:
        raise AssertionError(name)
    print("PASS", name)


def governed_manifest(source):
    return {
        "status": "READY",
        "source": {"filename": source, "sheet": "Datos", "row_count": 2, "source_fingerprint_sha256": "c" * 64},
        "components": [
            {"component_id": "metric:progress", "title": "Progreso", "status": "SUPPORTED", "value": 12, "unit": "count", "format": "integer", "provenance_source": "source"},
            {"component_id": "metric:service_level", "title": "Nivel de servicio", "status": "DERIVABLE", "value": 0.25, "unit": "ratio", "format": "percentage", "provenance_source": "approved-rule"},
            {"component_id": "metric:risk", "title": "Riesgo", "status": "BLOCKED", "value": 100, "unit": "score", "reason": "No hay regla aprobada", "provenance_source": "policy"},
            {"component_id": "metric:capacity", "title": "Capacidad", "status": "UNRESOLVED", "value": 200, "unit": "units", "reason": "Evidencia insuficiente", "provenance_source": "source"},
            {"component_id": "metric:definition", "title": "Definición", "status": "CONFLICT", "value": 300, "unit": "text", "reason": "Fuentes incompatibles", "provenance_source": "knowledge"},
        ],
    }


def html_manifest(path):
    html = path.read_text(encoding="utf-8")
    start = html.index("const DATA=") + len("const DATA=")
    end = html.index(";\n(()=>", start)
    return json.loads(html[start:end])["plan"]["enterprise_deliverable_manifest"], html


def excel_components(path):
    book = openpyxl.load_workbook(path, data_only=False)
    sheet = book["Resumen"]
    headers = [cell.value for cell in sheet[4]]
    rows = [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=5, values_only=True) if values[0]]
    metadata = {row[0]: row[1] for row in book["Metadata"].iter_rows(min_row=5, values_only=True) if row[0]}
    formulas = [cell.coordinate for ws in book.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    book.close()
    return rows, metadata, formulas


def normalized(rows):
    empty = lambda value: None if value in (None, "") else value
    return {row["Component ID"]: (row["Status"], empty(row["Value"]), empty(row["Unit"]) or "", empty(row["Reason"]) or "", empty(row["Provenance"]) or "") for row in rows}


def manifest_rows(manifest):
    rows = []
    for component in manifest["components"]:
        status = component["status"]
        value = None if status in {"BLOCKED", "UNRESOLVED", "CONFLICT"} else component["value"]
        rows.append({"Component ID": component["component_id"], "Status": status, "Value": value, "Unit": component["unit"], "Reason": component.get("reason") or "", "Provenance": component.get("provenance_source") or ""})
    return rows


def run_fixture(label, dataframe):
    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        source = f"{label}.csv"
        manifest = governed_manifest(source)
        profile = {"archivo": source, "filas": len(dataframe), "columnas": list(dataframe.columns), "run_id": "run-r10-20a-5", "roles_detectados": {}, "deliverable_manifest": manifest}
        sections = {"KPIs_Generales": pd.DataFrame([["Filas", len(dataframe)], ["Columnas", len(dataframe.columns)]], columns=["Indicador", "Valor"]), "Resultado": dataframe}
        plan = build_dashboard_plan(dataframe, "Resumen gobernado", source, "Datos")
        plan["enterprise_deliverable_manifest"] = manifest
        html_path, xlsx_path, pdf_path = root / "deliverable.html", root / "deliverable.xlsx", root / "deliverable.pdf"
        generate_dynamic_dashboard(html_path, dataframe, "Resumen gobernado", source, "Datos", prepared_plan=plan)
        professional.excel_report_professional(xlsx_path, "Resumen gobernado", profile, {"enterprise_deliverable_manifest": manifest}, sections, [], "", dataframe, dataframe, {}, "general")
        professional.pdf_report_professional(pdf_path, "Resumen gobernado", profile, sections, [], "", "general")
        html_evidence, html = html_manifest(html_path)
        excel_rows, metadata, formulas = excel_components(xlsx_path)
        pdf = fitz.open(pdf_path)
        pdf_text = "\n".join(page.get_text() for page in pdf)
        pdf.close()
        expected = normalized(manifest_rows(manifest))
        actual_excel = normalized(excel_rows)
        check(f"files_{label}", html_path.stat().st_size > 0 and xlsx_path.stat().st_size > 0 and pdf_path.read_bytes().startswith(b"%PDF"))
        check(f"same_html_evidence_{label}", html_evidence == manifest)
        check(f"excel_evidence_{label}", actual_excel == expected)
        check(f"pdf_ids_statuses_{label}", all(component_id in pdf_text and status in pdf_text for component_id, (status, *_rest) in expected.items()))
        check(f"pdf_values_{label}", "12" in pdf_text and "0.25" in pdf_text and "100" not in pdf_text and "200" not in pdf_text and "300" not in pdf_text)
        check(f"safety_{label}", not formulas and all(actual_excel[key][1] is None for key in ("metric:risk", "metric:capacity", "metric:definition")))
        check(f"metadata_{label}", metadata["source"] == source and metadata["design_system_version"] == DESIGN_SYSTEM_VERSION and "c" * 64 in pdf_text)
        regression_book = openpyxl.load_workbook(xlsx_path, read_only=True)
        regression_sheets = set(regression_book.sheetnames)
        regression_book.close()
        check(f"professional_regression_{label}", "dynamic-nav" in html and "--ds-primary" in html and {"Resumen", "Metadata", "Detalle"}.issubset(regression_sheets) and "Página 1" in pdf_text and "IA Empresarial Local" in pdf_text)


print("=== R10.20A.5 CROSS-OUTPUT ACCEPTANCE ===")
for name, frame in {
    "construccion": pd.DataFrame({"Proyecto": ["P-1", "P-2"], "Avance físico": [0.2, 0.4], "Presupuesto": [10, 20], "Estatus": ["Activo", "Pendiente"]}),
    "servicios": pd.DataFrame({"Contrato": ["C-1", "C-2"], "SLA": [0.9, 0.8], "Incidencias": [1, 2], "Tiempo atención": [2.5, 3.0]}),
    "logistica": pd.DataFrame({"Ruta": ["R-1", "R-2"], "Unidad operativa": ["U-1", "U-2"], "Carga": [8, 10], "Estado": ["En tránsito", "Completa"]}),
}.items():
    run_fixture(name, frame)
print("PASS R10.20A.5 CROSS-OUTPUT ACCEPTANCE")
