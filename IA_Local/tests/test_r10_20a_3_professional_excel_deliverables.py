from pathlib import Path
import sys
import tempfile

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

import reportes_profesionales as professional
from enterprise_design_system import DESIGN_SYSTEM_VERSION, get_excel_design_tokens


def check(name, condition):
    if not condition:
        raise AssertionError(name)
    print("PASS", name)


def manifest():
    return {
        "status": "READY",
        "source": {"filename": "source.csv", "sheet": "Datos", "row_count": 2, "source_fingerprint_sha256": "a" * 64},
        "components": [
            {"component_id": "kpi:advance", "title": "Avance", "status": "SUPPORTED", "value": 12, "unit": "count", "format": "integer", "provenance_source": "source"},
            {"component_id": "kpi:ratio", "title": "Cumplimiento", "status": "DERIVABLE", "value": 0.25, "unit": "ratio", "format": "percentage", "provenance_source": "rule"},
            {"component_id": "kpi:cost", "title": "Costo", "status": "BLOCKED", "value": 123, "reason": "No hay regla aprobada", "provenance_source": "policy"},
            {"component_id": "kpi:estimate", "title": "Estimación", "status": "UNRESOLVED", "value": 456, "reason": "Evidencia insuficiente"},
            {"component_id": "kpi:conflict", "title": "Conflicto", "status": "CONFLICT", "value": 789, "reason": "Fuentes incompatibles"},
        ],
    }


def build(domain, detail):
    profile = {"archivo": f"{domain}.csv", "filas": len(detail), "columnas": list(detail.columns), "run_id": "run-r10-20a-3", "roles_detectados": {}, "deliverable_manifest": manifest()}
    sections = {"KPIs_Generales": pd.DataFrame([["Filas", len(detail)], ["Columnas", len(detail.columns)]], columns=["Indicador", "Valor"]), "Resultado": detail}
    return profile, sections


print("=== R10.20A.3 PROFESSIONAL EXCEL DELIVERABLES ===")
check("canonical_generator", callable(professional.excel_report_professional))
tokens = get_excel_design_tokens()
check("design_system_excel_tokens", tokens["version"] == DESIGN_SYSTEM_VERSION and "SUPPORTED" in tokens["status"])

fixtures = {
    "construccion": pd.DataFrame({"Proyecto": ["P-01", "P-02"], "Avance físico": [0.25, 0.5], "Presupuesto": [1000.5, 2000.75], "Estatus": ["Activo", "En revisión"], "Fecha": pd.to_datetime(["2026-01-10", "2026-02-10"])}),
    "servicios": pd.DataFrame({"Contrato": ["C-01", "C-02"], "SLA": [0.95, 0.9], "Incidencias": [2, 3], "Tiempo atención": [1.25, 2.5], "Fecha": pd.to_datetime(["2026-03-10", "2026-04-10"])}),
    "logistica": pd.DataFrame({"Ruta": ["R-01", "R-02"], "Unidad operativa": ["U-1", "U-2"], "Carga": [10, 20], "Estado": ["Completa", "En tránsito"], "Fecha": pd.to_datetime(["2026-05-10", "2026-06-10"])}),
}

with tempfile.TemporaryDirectory() as td:
    for domain, detail in fixtures.items():
        path = Path(td) / f"{domain}.xlsx"
        profile, sections = build(domain, detail)
        professional.excel_report_professional(path, "Consulta UTF-8: áéíóú", profile, {"type": "overview"}, sections, [], "", detail, detail, {}, "general")
        check(f"xlsx_created_{domain}", path.is_file() and path.stat().st_size > 0)
        book = load_workbook(path, data_only=False)
        check(f"dynamic_sheets_{domain}", {"Dashboard", "Resumen", "Detalle", "Metadata"}.issubset(book.sheetnames))
        summary = book["Resumen"]
        detail_sheet = book["Detalle"]
        metadata = book["Metadata"]
        check(f"styles_{domain}", summary["A1"].font.bold and len(detail_sheet.tables) == 1)
        table = next(iter(detail_sheet.tables.values()))
        check(f"table_quality_{domain}", detail_sheet.freeze_panes == "A5" and table.autoFilter.ref is not None and detail_sheet.column_dimensions["A"].width <= 45)
        formats = [cell.number_format for sheet in (summary, detail_sheet) for row in sheet.iter_rows() for cell in row]
        check(f"number_date_formats_{domain}", "#,##0;[Red]-#,##0" in formats and "#,##0.00;[Red]-#,##0.00" in formats and "0.00%" in formats and any("yy" in value for value in formats))
        values = [cell.value for row in summary.iter_rows() for cell in row]
        check(f"governed_statuses_{domain}", all(status in values for status in ("SUPPORTED", "DERIVABLE", "BLOCKED", "UNRESOLVED", "CONFLICT")))
        blocked_row = next(row for row in summary.iter_rows(values_only=True) if row[2] == "BLOCKED")
        check(f"blocked_without_value_{domain}", blocked_row[3] is None and blocked_row[6] == "No hay regla aprobada")
        metadata_values = [cell.value for row in metadata.iter_rows() for cell in row]
        check(f"provenance_{domain}", DESIGN_SYSTEM_VERSION in metadata_values and "a" * 64 in metadata_values)
        check(f"no_secrets_{domain}", not any("password" in str(value).lower() or "connection string" in str(value).lower() for value in metadata_values))
        check(f"same_governed_values_{domain}", any(row[0] == "kpi:advance" and row[3] == 12 for row in summary.iter_rows(values_only=True)))
        formulas = [cell.coordinate for sheet in book.worksheets for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
        check(f"no_invented_formulas_{domain}", not formulas)
        book.close()

print("PASS R10.20A.3 PROFESSIONAL EXCEL DELIVERABLES")
